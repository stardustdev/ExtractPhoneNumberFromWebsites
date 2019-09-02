from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import load_workbook
import re

filepath = ("test.xlsx")        # This is file location.

wb = load_workbook(filepath)    # Load workbook and find active sheet.
sheet = wb.active

i = 1                           # This is line number of the first url existing.

while i > 0:

    # Take urls one by one
    i = i + 1
    url = sheet.cell(i,1).value
    if url == None:             # If url doesn't exist, iteration is finished
        break

    url = ('http://' + url, url) ['http://' in url]         # remake url
    print(url)

    # Take urls one by one. If request state is not 200 (success), then go to the next url
    try:
        page = urlopen(url)
        soup = BeautifulSoup(page, 'html.parser')
        innerhtml = soup.select("p")
        txt = ''
        for idx in innerhtml:       # parse innerhtml
            txt += '\n' + idx.text
        phones = re.findall(r'\(?[0-9]{3}\)?[ .-]?[0-9]{3}[ .-]?[0-9]{4}', txt) # parse phone numbers from the innerhtml

        # If on phone numbers in index page, then parse in contact page
        if phones == []:
            url += '/contact'
            print(url)

            try:
                page = urlopen(url)
                soup = BeautifulSoup(page, 'html.parser')
                innerhtml = soup.select("p")
                txt = ''
                for idx in innerhtml:
                    txt += '\n' + idx.text
                phones = re.findall(r'\(?[0-9]{3}\)?[ .-]?[0-9]{3}[ .-]?[0-9]{4}', txt)
            except:
                continue

    except:
        continue

    if (phones == []):
        continue

    phones = ', '.join(list(dict.fromkeys(phones)))     # remove duplicate phone numbers
    sheet.cell(i, 12).value = phones
    print(phones)

wb.save(filepath)



